#Tops To Template
import numpy as np
import openpyxl as xl
import urllib.request 
import requests

path ="C:/code-python/MaliCodes/ABaiMarLishop01/Upload-Malishop01-Lotus-Facefoam73.xlsx"

wbOrg = xl.load_workbook(path)
wsOrg = wbOrg.worksheets[1]

workpath = "C:/code-python/MaliCodes/ABaiMarLishop01/Images/Lotus-Facefoam"
folder = path[-1*(len(path)-path.rfind('/')) : ]
# filename = folder[1:folder.rfind('.')] + '.jpg'
# newPath = workpath + '/' + filename 

# print('New Path',newPath,'Filename : ', filename)

headOrg_row = 2
read_start_row = 3
# headTemp_row = 1
# write_start_row = 5

colOrg = [cell.value for cell in wsOrg[headOrg_row]]
# colTemp = [cell.value for cell in wsTemp[headTemp_row]]
# print(colOrg)
imageName = ['*รูปของสินค้า1','รูปของสินค้า2','รูปของสินค้า3','รูปของสินค้า4','รูปของสินค้า5',
             'รูปของสินค้า6','รูปของสินค้า7','รูปของสินค้า8']
noImage = len(imageName)-1
maxOrg_row = wsOrg.max_row - read_start_row + 1
headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1700.107 Safari/537.36' }
for i in range(0, maxOrg_row):
    order = 0
    Org_content = wsOrg.cell(row=read_start_row+i, column=colOrg.index(imageName[order])+1)
    sellerSKU = wsOrg.cell(row=read_start_row+i, column=colOrg.index('SellerSKU')+1)

    urlImage = (Org_content.value)
    while True:          
        ext = 'JPG'
        newFile = workpath+'/'+ str(sellerSKU.value)+'-'+str(order)+'.'+ext
        print('New file :',newFile, 'url Image : ', urlImage)

        urllib.request.urlretrieve(urlImage, newFile)
        urlImage = (wsOrg.cell(row=read_start_row+i, column=colOrg.index(imageName[order])+1).value)
        if (urlImage is None):
            break
        else:
            if (order < noImage):
                order = order+1
            else:
                break
            print(urlImage)