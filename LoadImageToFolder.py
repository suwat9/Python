#Tops To Template
import numpy as np
import openpyxl as xl
import urllib.request 
import requests

path ="D:/Codes/Mali-Python/Create Shop/Shop18/Airfresh/Malishop18-Upload-Thaiwatsadu-AirgoodSmell184.xlsx"
wbOrg = xl.load_workbook(path)
wsOrg = wbOrg.worksheets[1]

workpath = 'D:/Codes/Mali-Python/Create Shop/LoadImages/AirgoodSmell'
folder = path[-1*(len(path)-path.rfind('-')-1) : ]
filename = folder[0:folder.rfind('.')-1]
newPath = workpath + '/' + filename

headOrg_row = 2
read_start_row = 3
# headTemp_row = 1
# write_start_row = 5

colOrg = [cell.value for cell in wsOrg[headOrg_row]]
# colTemp = [cell.value for cell in wsTemp[headTemp_row]]
# print(colOrg)
imageName = ['*รูปของสินค้า1','รูปของสินค้า2','รูปของสินค้า3','รูปของสินค้า4','รูปของสินค้า5',
             'รูปของสินค้า6','รูปของสินค้า7','รูปของสินค้า8']
maxOrg_row = wsOrg.max_row - read_start_row + 1
headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1700.107 Safari/537.36' }
for i in range(0, maxOrg_row+1):
    order = 0
    Org_content = wsOrg.cell(row=read_start_row+i, column=colOrg.index(imageName[order])+1)
    sellerSKU = wsOrg.cell(row=read_start_row+i, column=colOrg.index('SellerSKU')+1)

    urlImage = str(Org_content.value)
    while (not urlImage == None):          
        print(urlImage)
        order = order + 1
        ext = urlImage[-1*(len(urlImage)-urlImage.rfind('.')-1) : ]
        newFile = newPath+'-'+ str(sellerSKU.value)+'-'+str(order)+'.'+ext

        urllib.request.urlretrieve(urlImage, newFile)
        urlImage = str(wsOrg.cell(row=read_start_row+i, column=colOrg.index(imageName[order])+1).value)