#Tops To Template
import numpy as np
import openpyxl as xl

fileTemplate =["C:/code-python/MaliCodes/Create Shop/editFields/Upload-Malishop01-Lotus-Softener304.xlsx",
               "C:/code-python/MaliCodes/Create Shop/editFields/Upload-Malishop05-Tops-Softener141Sku.xlsx",
               "C:/code-python/MaliCodes/Create Shop/editFields/Upload-Malishop11-BigC-Softener113.xlsx"]

for file in fileTemplate:
    wbTemp = xl.load_workbook(file)
    wsTemp = wbTemp.worksheets[1]

    headTemp_row = 2
    read_start_row = 3
    
    colTemp = [cell.value for cell in wsTemp[headTemp_row]]
    maxTemp_row = wsTemp.max_row - headTemp_row 
    
    
    for i in range(0, maxTemp_row):
        
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice Start')+1,value='2022-01-01')
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice End')+1,value='2026-12-31')
        order = str(i + 1)
        #shopName = 'Malishop14-Top-Rice'+order.zfill(4)
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('SellerSKU')+1,value=shopName)
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ยี่ห้อ')+1,value='Premium(พรีเมึ่ยม)')
        imgLink = wsTemp.cell(row=read_start_row+i, column=colTemp.index('*รูปของสินค้า1')+1)
        detail = wsTemp.cell(row=read_start_row+i, column=colTemp.index('Long Description (Lorikeet)')+1)
        #if detail.value==None:
        #    detail = wsOrg.cell(row=read_start_row+i, column=colOrg.index('ProductName2')+1)
        content = "<span style=\"font-family:none\">"+ str(detail.value) + "<\span><div style=\"width:100%;margin:0\"><div style=\"width:100%;display:block;margin:0\"><div style=\"width:100%;display:block\"><img class=\"\" src=\""+ str(imgLink.value) +"\" style=\"width:100%;display:block\"\></div></div></div>"
        wsTemp.cell(row=read_start_row+i, column=colTemp.index('Long Description (Lorikeet)')+1,value=content)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*น้ำหนัก')+1,value=1)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความยาว')+1,value=17)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความกว้าง')+1,value=11)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความสูง')+1,value=1)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*จำนวน')+1,value=10)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('รุ่น')+1,value='ไม่ระบุ')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('การประกัน')+1,value='1 สัปดาห์')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*การรับประกัน')+1,value='มีการรับประกัน')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('สินค้าอันตราย')+1,value='ไม่มี')
    wbTemp.save(file)
    wbTemp.close()
