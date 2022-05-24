#Tops To Template
import numpy as np
import openpyxl as xl
from os import walk
import os
from random import seed,randint

path = 'D:/Codes/Mali-Python/Create Shop/editFields/Thaiwatsadu'
workPath = 'D:/Codes/Mali-Python/Create Shop/Shop16'
folder = path[-1*(len(path)-path.rfind('/')-1) : ]
newPath = workPath + '/' + folder 
if not os.path.isdir(os.path.join(os.getcwd(), newPath)):
    os.mkdir(newPath)
    
fileTemplate = next(walk(path), (None, None, []))[2]  # [] if no file

#if not os.path.isdir(os.path.join(os.getcwd(), splitPath)):
#    os.mkdir(splitPath)
#print(fileTemplate)

for filename in fileTemplate:
    file = path + '/' + filename
    shopName = 'Malishop16'
    newFile = newPath + '/'+shopName + '-' + filename
    wbTemp = xl.load_workbook(file)
    wsTemp = wbTemp.worksheets[1]

    headTemp_row = 2
    read_start_row = 3
    
    colTemp = [cell.value for cell in wsTemp[headTemp_row]]
    maxTemp_row = wsTemp.max_row - headTemp_row 
    # seed number for discount random from 0 to 10 bahts
    seednumber = 3
    seed(seednumber)

    
    for i in range(0, maxTemp_row):
        
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice Start')+1,value='2022-01-01')
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice End')+1,value='2026-12-31')
        order = str(i + 1)
        # shopName = 'Malishop16-Top-Rice'+order.zfill(4)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('SellerSKU')+1,value=shopName)
        #wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ยี่ห้อ')+1,value='Premium(พรีเมึ่ยม)')
        
        # insert HTML to Long Description (Lorikeet)
        imgLink = wsTemp.cell(row=read_start_row+i, column=colTemp.index('*รูปของสินค้า1')+1)
        detail = wsTemp.cell(row=read_start_row+i, column=colTemp.index('Long Description (Lorikeet)')+1)
        content = "<span style=\"font-family:none\">"+ str(detail.value) + "<\span><div style=\"width:100%;margin:0\"><div style=\"width:100%;display:block;margin:0\"><div style=\"width:100%;display:block\"><img class=\"\" src=\""+ str(imgLink.value) +"\" style=\"width:100%;display:block\"\></div></div></div>"
        wsTemp.cell(row=read_start_row+i, column=colTemp.index('Long Description (Lorikeet)')+1,value=content)
        
        # Change keyword in 3 fields
        keyName = ['*ชื่อสินค้า','Long Description (Lorikeet)','สินค้าภายในกล่อง']
        productName = wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[0])+1)
        longDesc = wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[1])+1)
        productBox = wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[2])+1)
        keyText = ['จัดส่งฟรี ','ส่งฟรี ','ฟรี!จัดส่ง', 'ส่งฟรี*** ','ส่งฟรีครับผม  ','ส่งฟรีส่งเร็ว  ']
        promoteWord = 'ฟรีค่าส่ง '
        newText = [promoteWord for cell in keyText]
        #newText = [promoteWord, promoteWord, promoteWord, promoteWord, promoteWord, promoteWord]
        found = -1
        ind = 0
        
        while (found<0) and (ind<len(keyText)):
            found = str(productName.value).find(keyText[ind])
            if found>=0: keyind = ind 
            else: keynotfound = ind
            if ind < len(keyText): ind = ind+1
                
        if found>=0:
            keyValue = [str(productName.value).replace(keyText[keyind], newText[keyind]),
                        str(longDesc.value).replace(keyText[keyind], newText[keyind]),
                        str(productBox.value).replace(keyText[keyind], newText[keyind])]
            wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[0])+1,value=keyValue[0])
            wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[1])+1,value=keyValue[1])
            wsTemp.cell(row=read_start_row+i, column=colTemp.index(keyName[2])+1,value=keyValue[2])
        else:
            if i==0: print('File : '+filename+' ไม่พบคำว่า '+keyText[keynotfound])

        #random discount 0-10 bahts
        discount = randint(0, 10)
        spPrice = wsTemp.cell(row=read_start_row+i, column=colTemp.index('SpecialPrice')+1)
        newPrice = int(spPrice.value)-discount
        wsTemp.cell(row=read_start_row+i, column=colTemp.index('SpecialPrice')+1,value=newPrice)
        
        oldShopName = wsTemp.cell(row=read_start_row+i, column=colTemp.index('SellerSKU')+1)
        
        wsTemp.cell(row=read_start_row+i, column=colTemp.index('SellerSKU')+1,value=shopName+str(oldShopName.value)[10:])
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*น้ำหนัก')+1,value=1)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความยาว')+1,value=17)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความกว้าง')+1,value=11)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความสูง')+1,value=1)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*จำนวน')+1,value=10)
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('รุ่น')+1,value='ไม่ระบุ')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('การประกัน')+1,value='1 สัปดาห์')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('*การรับประกัน')+1,value='มีการรับประกัน')
        # wsTemp.cell(row=write_start_row+i, column=colTemp.index('สินค้าอันตราย')+1,value='ไม่มี')
    wbTemp.save(newFile)
    wbTemp.close()