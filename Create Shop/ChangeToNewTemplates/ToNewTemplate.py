#Tops To Template
import numpy as np
import openpyxl as xl

fileOrg ="D:/Codes/Mali-Python/Create Shop/ChangeToNewTemplates/Old-Template/Upload-TopEng-Tissue169.xlsx"
wbOrg = xl.load_workbook(fileOrg)
wsOrg = wbOrg.worksheets[1]

fileTemplate ="D:/Codes/Mali-Python/Create Shop/ChangeToNewTemplates/NewFiles/New-Upload-TopEng-Tissue169.xlsx"
wbTemp = xl.load_workbook(fileTemplate)
wsTemp = wbTemp.worksheets[1]

headOrg_row = 2
read_start_row = 3
headTemp_row = 1
write_start_row = 5

colOrg = [cell.value for cell in wsOrg[headOrg_row]]
colTemp = [cell.value for cell in wsTemp[headTemp_row]]
maxOrg_row = wsOrg.max_row - read_start_row + 1

if len(colOrg)==len(colTemp):
    print('OK: Columns matching.\n')
    col_list = np.dstack((colOrg,colTemp))[0]
    print(col_list)
    # check columns before doing something
    ans = 'Y'
    # ans = input('Column is OK [y/n]:')
    if (ans in ['y','Y']):
        for i in range(0, maxOrg_row+1):
            for c in range(0,col_list.shape[0]):
                Org_content = wsOrg.cell(row=read_start_row+i, column=c+1)               
                wsTemp.cell(row=write_start_row+i, column=c+1, value=Org_content.value)
        wbTemp.save(fileTemplate)
        wbTemp.close()
        wbOrg.close()
    else:
        print('Quit.......')

else:
    print('Old template:\n',colOrg)
    print('\nNew Template:\n',colTemp)

