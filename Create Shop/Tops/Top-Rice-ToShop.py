#Tops To Template
import numpy as np
import openpyxl as xl

fileOrg ="D:/Codes/Mali-Python/Create Shop/Tops/Split-Eng - Tops-Rice - JAN2022 - 130SKU.xlsx"
wbOrg = xl.load_workbook(fileOrg)
wsOrg = wbOrg.worksheets[0]

fileTemplate ="D:/Codes/Mali-Python/Create Shop/shop14/Upload-Malishop14-TopEng-Rice132.xlsx"
wbTemp = xl.load_workbook(fileTemplate)
wsTemp = wbTemp.worksheets[1]

headOrg_row = 1
headTemp_row = 2
read_start_row = 2
write_start_row = 3

colOrg = [cell.value for cell in wsOrg[headOrg_row]]
colTemp = [cell.value for cell in wsTemp[headTemp_row]]
maxOrg_row = wsOrg.max_row - headOrg_row
maxTemp_row = wsTemp.max_row - headTemp_row

col_list = [['ProductName2','*ชื่อสินค้า'],['Pic1-src','*รูปของสินค้า1'],
            ['Pic1-src','Showcase_image1:1'],['ProductName2','สินค้าภายในกล่อง'],
            ['Pic2-src','รูปของสินค้า2'],['Pic3-src','รูปของสินค้า3'],
            ['Pic4-src','รูปของสินค้า4'],['Pic5-src','รูปของสินค้า5'],
            ['ProductName2','*ชื่อสินค้า'],['Pic1-src','รูปภาพ1'],
            ['Pic2-src','รูปภาพ2'],['Pic3-src','รูปภาพ3'],
            ['Pic4-src','รูปภาพ4'],['Pic5-src','รูปภาพ5'],
            ['PriceX','*ราคา'],['SpecialPrice','SpecialPrice']            
            ]

colPair = np.array(col_list)

for i in range(0, maxOrg_row+1):
    for c in range(0,colPair.shape[0]):
        #copy columns
        colOrg_index = colOrg.index(col_list[c][0])
        colTemp_index = colTemp.index(col_list[c][1])
        if ((colOrg_index >=0) and (colTemp_index >=0)):
            #copy from wb1
            content = wsOrg.cell(row=read_start_row+i, column=colOrg_index+1)
            #paste in ws2
            wsTemp.cell(row=write_start_row+i, column=colTemp_index+1, value=content.value)
        
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice Start')+1,value='2022-01-01')
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('SpecialPrice End')+1,value='2026-12-31')
    order = str(i + 1)
    shopName = 'Malishop14-Top-Rice'+order.zfill(4)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('SellerSKU')+1,value=shopName)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ยี่ห้อ')+1,value='Premium(พรีเมึ่ยม)')
    imgLink = wsOrg.cell(row=read_start_row+i, column=colOrg.index('Pic1-src')+1)
    detail = wsOrg.cell(row=read_start_row+i, column=colOrg.index('Description')+1)
    if detail.value==None:
        detail = wsOrg.cell(row=read_start_row+i, column=colOrg.index('ProductName2')+1)
    content = "<span style=\"font-family:none\">"+ str(detail.value) + "</span><div style=\"width:100%;margin:0\"><div style=\"width:100%;display:block;margin:0\"><div style=\"width:100%;display:block\"><img class=\"\" src=\""+ str(imgLink.value) +"\" style=\"width:100%;display:block\"/></div></div></div>"
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('Long Description (Lorikeet)')+1,value=content)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*น้ำหนัก')+1,value=1)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความยาว')+1,value=17)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความกว้าง')+1,value=11)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*ความสูง')+1,value=1)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*จำนวน')+1,value=10)
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('รุ่น')+1,value='ไม่ระบุ')
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('การประกัน')+1,value='1 สัปดาห์')
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('*การรับประกัน')+1,value='มีการรับประกัน')
    wsTemp.cell(row=write_start_row+i, column=colTemp.index('สินค้าอันตราย')+1,value='ไม่มี')
wbTemp.save(fileTemplate)
wbTemp.close()
wbOrg.close()